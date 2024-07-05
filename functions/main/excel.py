from openpyxl.styles import NamedStyle, PatternFill, Border, Side
from openpyxl import load_workbook
from pathlib import Path
import shutil
import unicodedata
from pathlib import Path
from datetime import datetime
import configparser
import functions as fnc

class ExcelWorkbook:
    def __init__(self, file_path, table_list_dir):
        self.root_path = file_path
        self.table_list_dir = table_list_dir  
        self.file_path = ''
        self.workbook = None
        self.sheet_names = []
        self.column_names = []
        self.soft_green_with_border = None  # Store the named style object

    def copy_excel_file(self):
        today_date = datetime.today().strftime('%Y_%m_%d_%H_%M_%S')

        # Get the current working directory
        current_dir = Path.cwd()

        # Create a new directory for storing Excel files if it doesn't exist
        excel_folder = current_dir / 'excel'
        if not excel_folder.exists():
            excel_folder.mkdir()

        # Copy the Excel file to the excel_folder with timestamp suffix
        destination_file_path = excel_folder / (self.root_path.stem + today_date + self.root_path.suffix)
        try:
            shutil.copy(self.root_path, destination_file_path)
            print(f"Copied {self.root_path} to {destination_file_path}")
            return destination_file_path  # Return the copied file path
        except PermissionError as e:
            print(f"Permission error: {e}")
        except Exception as e:
            print(f"Error copying {self.file_path}: {e}")
        
        return None    

    def load_workbook(self, copy_file_path, read_only=True):
        """Load the Excel workbook."""
        self.file_path = copy_file_path
        self.workbook = load_workbook(self.file_path, read_only=read_only)
        self.sheet_names = self.workbook.sheetnames

    def create_named_styles(self):
        """Create or retrieve named styles for cell formatting."""
        if not hasattr(self, 'soft_green_with_border') or not self.workbook.named_styles or 'soft_green_with_border' not in self.workbook.named_styles:
            # Only create if it doesn't exist
            self.soft_green_with_border = NamedStyle(name='soft_green_with_border')
            self.soft_green_with_border.fill = PatternFill(start_color='C0FFC0', end_color='C0FFC0', fill_type='solid')
            self.soft_green_with_border.border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))
            self.workbook.add_named_style(self.soft_green_with_border)
        else:
            # Modify existing style if needed
            self.soft_green_with_border.fill = PatternFill(start_color='C0FFC0', end_color='C0FFC0', fill_type='solid')
            self.soft_green_with_border.border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))

    def read_sheet_data(self, sheet):
        """Read data from a sheet and structure it into the desired format."""
        self.create_named_styles()

        if not sheet.protection.sheet:
            for col in range(2, sheet.max_column + 1): 
                main_cell = sheet.cell(row=2, column=col)
                if main_cell.value is not None:
                    table_name = main_cell.value
                    table_name = table_name.replace('[', '')
                    table_name = table_name.replace(']', '').lower()
                    self.extrate_column_names(table_name, sheet, col)
                    #main_cell.value is table name 
                    #if main_cell.value is exit in log dir, read data and convert array and store in self.column_names
                    
                    

            # Save workbook after making changes
            self.workbook.save(self.file_path)
        else:
            print(f"Sheet '{sheet.title}' is protected and cannot be modified.")

    def print_sheet_names_and_data(self):
        """Print sheet names and read structured data, skipping the first sheet."""
        for index, sheet_name in enumerate(self.sheet_names):
            if index != 0:
                sheet = self.workbook[sheet_name]
                self.read_sheet_data(sheet)

    def extrate_column_names(self, table_name, sheet, col):
        files = [file for file in self.table_list_dir.rglob('*') if file.is_file()]
        columns = []
        for file in files:
            file_name = file.name.replace('.txt','').lower()
            table_name = unicodedata.normalize('NFKC', table_name)
            file_name = unicodedata.normalize('NFKC', file_name)

            if file_name == table_name:    
                columns = fnc.read_file_to_array(file)
                if columns != []:
                    for row in range(3, sheet.max_row + 1):
                        child_cell = sheet.cell(row=row, column=col)
                        child_cell_value = child_cell.value
                        if child_cell_value is not None:  # If the child cell has data
                            #if child_cell.value exit in self.column_names
                            child_cell_value = unicodedata.normalize('NFKC', child_cell_value)
                            for column in columns:
                                if child_cell_value.lower() == column:
                                    child_cell.style = 'soft_green_with_border'

    
def process_excel_file():
    config = configparser.ConfigParser()
    current_root = Path().resolve()
    config.read('config.ini', encoding='utf-8')
    excelFiledir = Path(config['DEFAULT']['excel_file_dir'])
    tableListFiledir = current_root / Path(config['DEFAULT']['table_list_file_dir'])
    excel_workbook = ExcelWorkbook(Path(excelFiledir), Path(tableListFiledir))
    copied_file_path = excel_workbook.copy_excel_file()
    
    if copied_file_path:
        excel_workbook.load_workbook(copied_file_path, read_only=False)  # Load workbook in editable mode
        excel_workbook.print_sheet_names_and_data()
    else:
        print("Copying failed.")

if __name__ == "__main__":
    process_excel_file()


